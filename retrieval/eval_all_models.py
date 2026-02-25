"""
Đánh giá Retrieval cho nhiều model + Reranking
==============================================
Tác giả: Auto-generated
Ngày: 2025

Công dụng:
  - Evaluate retrieval performance của nhiều model (tfidf, legalhf, phobert, dek21)
  - Áp dụng reranking (cross-encoder)
  - Xuất kết quả ra Excel để so sánh

Usage:
  python eval_all_models.py --gold retrieval/gold_200_diverse.jsonl --chunks output_nghidinh/chunks_clean_norm.json --output eval_results.xlsx

Metrics:
  - Recall@K (K=1,3,5,10)
  - MRR (Mean Reciprocal Rank)
  - Hit@K
"""

import os
import json
import argparse
import time
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field

import numpy as np

# Dense retrieval
import faiss
from sentence_transformers import SentenceTransformer

# TF-IDF
import joblib
import scipy.sparse as sp
from sklearn.metrics.pairwise import linear_kernel

# Reranking (cross-encoder)
try:
    from sentence_transformers import CrossEncoder
    HAS_CROSS_ENCODER = True
except ImportError:
    HAS_CROSS_ENCODER = False

# Excel export
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_PANDAS_OPENPYXL = True
except ImportError:
    HAS_PANDAS_OPENPYXL = False


# =============================================================================
# DATA CLASSES
# =============================================================================
@dataclass
class EvalConfig:
    """Cấu hình đánh giá"""
    models: List[str] = field(default_factory=lambda: ["tfidf", "legalhf", "phobert", "dek21"])
    topk: int = 100  # Retrieve top-100 trước khi rerank
    match_mode: str = "dieu_khoan"  # "dieu" hoặc "dieu_khoan"
    use_rerank: bool = True
    rerank_model: str = "BAAI/bge-reranker-v2-m3"  # Multilingual, tốt cho tiếng Việt
    rerank_topk: int = 10  # Sau rerank lấy top-10
    device: str = "cuda"
    vec_root: str = "vector_data"
    chunks_path: str = "output_nghidinh/chunks_clean_norm.json"
    save_detailed: bool = True  # Ghi log chi tiết từng query


@dataclass
class RetrievalResult:
    """Kết quả retrieval cho 1 query"""
    query: str
    qid: str
    model_name: str
    topk_results: List[Dict]  # [{rank, score, text, metadata}]
    ground_truth: Dict
    found_rank: int  # 0 if not found in topk


@dataclass
class ModelMetrics:
    """Metrics cho 1 model"""
    model_name: str
    reranked: bool
    n_queries: int
    recall_at: Dict[int, float]  # {1: 0.xx, 3: 0.xx, ...}
    mrr: float
    hit_at: Dict[int, int]  # raw counts
    avg_time_per_query: float
    timestamp: str


# =============================================================================
# HELPERS
# =============================================================================
def load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def load_jsonl(path: str) -> List[Dict]:
    items = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                items.append(json.loads(line))
    return items

def norm_str(x) -> str:
    if x is None:
        return ""
    return str(x).strip().lower()

def match_ground_truth(gt: Dict, meta: Dict, mode: str = "dieu_khoan") -> bool:
    """
    So khớp ground truth với metadata của chunk
    mode:
      - "dieu": khớp van_ban + dieu
      - "dieu_khoan": khớp van_ban + dieu + khoan
    """
    # Lấy van_ban từ gt - có thể có key khác nhau
    gt_van_ban = gt.get("van_ban", "") or gt.get("source_contains", "")
    meta_van_ban = meta.get("van_ban", "")
    
    # Normalize và so sánh (có thể dùng contains cho linh hoạt)
    gt_vb_norm = norm_str(gt_van_ban)
    meta_vb_norm = norm_str(meta_van_ban)
    
    # Kiểm tra van_ban - dùng contains vì tên văn bản có thể khác nhau
    vb_match = False
    if gt_vb_norm and meta_vb_norm:
        # Thử exact match trước
        if gt_vb_norm == meta_vb_norm:
            vb_match = True
        # Thử contains
        elif gt_vb_norm in meta_vb_norm or meta_vb_norm in gt_vb_norm:
            vb_match = True
    elif not gt_vb_norm:
        # Nếu gt không có van_ban thì bỏ qua điều kiện này
        vb_match = True
    
    if not vb_match:
        return False
    
    # Kiểm tra điều
    if norm_str(gt.get("dieu")) != norm_str(meta.get("dieu")):
        return False
    
    if mode == "dieu":
        return True
    
    # Kiểm tra khoản
    gt_khoan = gt.get("khoan")
    meta_khoan = meta.get("khoan")
    
    # Nếu gt không có khoản (None) thì coi như match
    if gt_khoan is None:
        return True
    
    return norm_str(gt_khoan) == norm_str(meta_khoan)


def mrr_from_ranks(ranks: List[int]) -> float:
    """Tính MRR từ list các rank (1-based), 0 nếu không tìm thấy"""
    if not ranks:
        return 0.0
    s = sum(1.0/r if r > 0 else 0.0 for r in ranks)
    return s / len(ranks)


def l2_normalize(x: np.ndarray, eps: float = 1e-12) -> np.ndarray:
    n = np.linalg.norm(x, axis=1, keepdims=True)
    return x / np.clip(n, eps, None)


# =============================================================================
# MODEL LOADERS
# =============================================================================
MODEL_MAP = {
    "legalhf": {
        "dir": "legal_hf_cosine",
        "encoder": "Quockhanh05/Vietnam_legal_embeddings",
        "type": "dense",
        "normalize_query": True,
    },
    "phobert": {
        "dir": "phobert", 
        "encoder": "vinai/phobert-base",
        "type": "dense",
        "normalize_query": True,
    },
    "dek21": {
        "dir": "dek21",
        "encoder": "dangvantuan/vietnamese-embedding",
        "type": "dense",
        "normalize_query": True,
    },
    "tfidf": {
        "dir": "tfidf",
        "type": "tfidf",
    }
}


class ModelCache:
    """Cache các model đã load để tránh load lại"""
    
    def __init__(self, vec_root: str, device: str = "cuda"):
        self.vec_root = vec_root
        self.device = device
        self._encoders = {}
        self._indexes = {}
        self._tfidf = {}
    
    def get_encoder(self, model_key: str) -> SentenceTransformer:
        if model_key not in self._encoders:
            model_name = MODEL_MAP[model_key]["encoder"]
            print(f"   Loading encoder: {model_name}...")
            self._encoders[model_key] = SentenceTransformer(model_name, device=self.device)
        return self._encoders[model_key]
    
    def get_faiss_index(self, model_key: str) -> faiss.Index:
        if model_key not in self._indexes:
            vec_dir = os.path.join(self.vec_root, MODEL_MAP[model_key]["dir"])
            idx_path = os.path.join(vec_dir, "index.faiss")
            if not os.path.isfile(idx_path):
                raise FileNotFoundError(f"Không tìm thấy {idx_path}")
            print(f"   Loading FAISS index: {idx_path}...")
            self._indexes[model_key] = faiss.read_index(idx_path)
        return self._indexes[model_key]
    
    def get_tfidf(self, model_key: str = "tfidf"):
        if model_key not in self._tfidf:
            vec_dir = os.path.join(self.vec_root, MODEL_MAP[model_key]["dir"])
            
            # Load vectorizer
            vec_path = os.path.join(vec_dir, "vectorizer.pkl")
            if not os.path.isfile(vec_path):
                vec_path = os.path.join(vec_dir, "vectorizer.joblib")
            if not os.path.isfile(vec_path):
                raise FileNotFoundError(f"Không tìm thấy vectorizer tại {vec_dir}")
            
            # Load matrix
            mat_path = os.path.join(vec_dir, "tfidf_matrix.npz")
            if not os.path.isfile(mat_path):
                raise FileNotFoundError(f"Không tìm thấy tfidf_matrix.npz tại {vec_dir}")
            
            print(f"   Loading TF-IDF: {vec_dir}...")
            vectorizer = joblib.load(vec_path)
            tfidf_matrix = sp.load_npz(mat_path)
            self._tfidf[model_key] = (vectorizer, tfidf_matrix)
        
        return self._tfidf[model_key]


# =============================================================================
# RETRIEVERS
# =============================================================================
def dense_retrieve(
    query: str,
    topk: int,
    encoder: SentenceTransformer,
    index: faiss.Index,
    chunks: List[Dict],
    normalize_query: bool = True,
) -> List[Dict]:
    """Retrieve với dense model (FAISS)"""
    q_vec = encoder.encode([query], convert_to_numpy=True).astype("float32")
    if normalize_query:
        q_vec = l2_normalize(q_vec)
    
    scores, ids = index.search(q_vec, topk)
    scores = scores[0]
    ids = ids[0]
    
    results = []
    for rank, (idx, sc) in enumerate(zip(ids, scores), start=1):
        if idx < 0 or idx >= len(chunks):
            continue
        c = chunks[int(idx)]
        results.append({
            "rank": rank,
            "score": float(sc),
            "text": c.get("text", ""),
            "metadata": c.get("metadata", {}),
            "chunk_idx": int(idx),
        })
    return results


def tfidf_retrieve(
    query: str,
    topk: int,
    vectorizer,
    tfidf_matrix,
    chunks: List[Dict],
) -> List[Dict]:
    """Retrieve với TF-IDF"""
    q_vec = vectorizer.transform([query])
    sims = linear_kernel(q_vec, tfidf_matrix).ravel()
    
    if topk >= len(sims):
        top_idx = np.argsort(-sims)
    else:
        top_idx = np.argpartition(-sims, topk)[:topk]
        top_idx = top_idx[np.argsort(-sims[top_idx])]
    
    results = []
    for rank, idx in enumerate(top_idx[:topk], start=1):
        c = chunks[int(idx)]
        results.append({
            "rank": rank,
            "score": float(sims[int(idx)]),
            "text": c.get("text", ""),
            "metadata": c.get("metadata", {}),
            "chunk_idx": int(idx),
        })
    return results


def retrieve(
    query: str,
    topk: int,
    model_key: str,
    chunks: List[Dict],
    model_cache: ModelCache,
) -> List[Dict]:
    """Unified retrieve function"""
    model_config = MODEL_MAP[model_key]
    
    if model_config["type"] == "tfidf":
        vectorizer, tfidf_matrix = model_cache.get_tfidf(model_key)
        return tfidf_retrieve(query, topk, vectorizer, tfidf_matrix, chunks)
    else:
        encoder = model_cache.get_encoder(model_key)
        index = model_cache.get_faiss_index(model_key)
        normalize = model_config.get("normalize_query", True)
        return dense_retrieve(query, topk, encoder, index, chunks, normalize)


# =============================================================================
# RERANKING
# =============================================================================
class Reranker:
    """Cross-encoder reranker"""
    
    def __init__(self, model_name: str = "BAAI/bge-reranker-base", device: str = "cuda"):
        if not HAS_CROSS_ENCODER:
            raise ImportError("CrossEncoder không khả dụng. Cài: pip install sentence-transformers")
        
        print(f"Loading reranker: {model_name}...")
        self.model = CrossEncoder(model_name, device=device)
        self.model_name = model_name
    
    def rerank(
        self,
        query: str,
        candidates: List[Dict],
        topk: int = 5,
    ) -> List[Dict]:
        """Rerank candidates và trả về top-k"""
        if not candidates:
            return []
        
        # Chuẩn bị pairs
        pairs = [(query, c.get("text", "")) for c in candidates]
        
        # Score
        scores = self.model.predict(pairs)
        
        # Kết hợp score vào results
        scored = []
        for i, c in enumerate(candidates):
            scored.append({
                **c,
                "original_score": c["score"],
                "rerank_score": float(scores[i]),
            })
        
        # Sort theo rerank_score
        scored.sort(key=lambda x: x["rerank_score"], reverse=True)
        
        # Update rank và trả về top-k
        results = []
        for new_rank, item in enumerate(scored[:topk], start=1):
            item["rank"] = new_rank
            item["score"] = item["rerank_score"]  # Update score chính
            results.append(item)
        
        return results


# =============================================================================
# EVALUATOR
# =============================================================================
class RetrievalEvaluator:
    """Đánh giá retrieval cho nhiều models"""
    
    def __init__(self, config: EvalConfig):
        self.config = config
        self.chunks = load_json(config.chunks_path)
        self.model_cache = ModelCache(config.vec_root, config.device)
        self.reranker = None
        self.detailed_logs = {}  # {model_key: [query_details]}
        
        if config.use_rerank and HAS_CROSS_ENCODER:
            try:
                self.reranker = Reranker(config.rerank_model, config.device)
            except Exception as e:
                print(f"⚠️ Không load được reranker: {e}")
    
    def evaluate_model(
        self,
        model_key: str,
        gold_data: List[Dict],
        use_rerank: bool = False,
    ) -> ModelMetrics:
        """Đánh giá 1 model trên gold data"""
        
        log_key = f"{model_key}{'_rerank' if use_rerank else ''}"
        print(f"\n{'='*60}")
        print(f"Evaluating: {model_key} {'+ Rerank' if use_rerank else ''}")
        print(f"{'='*60}")
        
        Ks = [1, 3, 5, 10]
        Ks = [k for k in Ks if k <= self.config.topk]
        
        hit_at = {k: 0 for k in Ks}
        ranks = []
        total_time = 0.0
        query_details = []  # Lưu chi tiết từng query
        
        for i, g in enumerate(gold_data):
            query = g.get("question", "")
            qid = g.get("qid", f"Q{i+1}")
            
            # Lấy ground truth (có thể nằm trong list hoặc trực tiếp)
            gt_list = g.get("ground_truth", [g])  # Nếu không có ground_truth, dùng chính g
            if not isinstance(gt_list, list):
                gt_list = [gt_list]
            
            # Ground truth string để log
            gt_dieu = gt_list[0].get("dieu", "") if gt_list else ""
            gt_khoan = gt_list[0].get("khoan", "") if gt_list else ""
            gt_van_ban = gt_list[0].get("van_ban", "") if gt_list else ""
            
            # Retrieve
            start_time = time.time()
            
            try:
                results = retrieve(
                    query, 
                    self.config.topk, 
                    model_key, 
                    self.chunks, 
                    self.model_cache
                )
            except Exception as e:
                print(f"   ❌ Lỗi retrieve [{qid}]: {e}")
                ranks.append(0)
                query_details.append({
                    "qid": qid,
                    "question": query,
                    "gt_dieu": gt_dieu,
                    "gt_khoan": gt_khoan,
                    "gt_van_ban": gt_van_ban[:50] if gt_van_ban else "",
                    "found_rank": 0,
                    "is_hit": False,
                    "top1_dieu": "",
                    "top1_khoan": "",
                    "top1_score": 0,
                    "top1_text": "",
                    "error": str(e),
                })
                continue
            
            # Rerank nếu cần
            if use_rerank and self.reranker:
                results = self.reranker.rerank(
                    query, 
                    results, 
                    topk=self.config.rerank_topk
                )
            
            elapsed = time.time() - start_time
            total_time += elapsed
            
            # Tìm rank của kết quả đúng đầu tiên
            found_rank = 0
            for r in results:
                meta = r.get("metadata", {})
                for gt in gt_list:
                    if match_ground_truth(gt, meta, self.config.match_mode):
                        found_rank = r["rank"]
                        break
                if found_rank > 0:
                    break
            
            ranks.append(found_rank)
            
            # Update hit@k
            for k in Ks:
                if 0 < found_rank <= k:
                    hit_at[k] += 1
            
            # Lưu chi tiết query
            top1 = results[0] if results else {}
            top1_meta = top1.get("metadata", {})
            detail = {
                "qid": qid,
                "question": query[:100],  # Giới hạn độ dài
                "gt_dieu": gt_dieu,
                "gt_khoan": gt_khoan,
                "gt_van_ban": gt_van_ban[:50] if gt_van_ban else "",
                "found_rank": found_rank,
                "is_hit@1": found_rank == 1,
                "is_hit@5": 0 < found_rank <= 5,
                "is_hit@10": 0 < found_rank <= 10,
                "top1_dieu": top1_meta.get("dieu", ""),
                "top1_khoan": top1_meta.get("khoan", ""),
                "top1_van_ban": str(top1_meta.get("van_ban", ""))[:50],
                "top1_score": round(top1.get("score", 0), 4),
                "top1_text": top1.get("text", "")[:100],
                "time_ms": round(elapsed * 1000, 1),
            }
            # Thêm top-3 results
            for j in range(min(3, len(results))):
                r = results[j]
                m = r.get("metadata", {})
                detail[f"top{j+1}_dieu"] = m.get("dieu", "")
                detail[f"top{j+1}_khoan"] = m.get("khoan", "")
                detail[f"top{j+1}_score"] = round(r.get("score", 0), 4)
            
            query_details.append(detail)
            
            # Progress
            if (i + 1) % 50 == 0 or (i + 1) == len(gold_data):
                print(f"   Processed: {i+1}/{len(gold_data)}")
        
        # Lưu detailed logs
        self.detailed_logs[log_key] = query_details
        
        # Tính metrics
        n = len(gold_data)
        recall_at = {k: hit_at[k] / n if n > 0 else 0.0 for k in Ks}
        mrr = mrr_from_ranks(ranks)
        avg_time = total_time / n if n > 0 else 0.0
        
        metrics = ModelMetrics(
            model_name=model_key,
            reranked=use_rerank,
            n_queries=n,
            recall_at=recall_at,
            mrr=mrr,
            hit_at=hit_at,
            avg_time_per_query=avg_time,
            timestamp=datetime.now().isoformat(),
        )
        
        # In kết quả
        print(f"\n   Results for {model_key} {'+ Rerank' if use_rerank else ''}:")
        for k in Ks:
            print(f"      Recall@{k}: {recall_at[k]:.4f} ({hit_at[k]}/{n})")
        print(f"      MRR: {mrr:.4f}")
        print(f"      Avg time/query: {avg_time*1000:.1f}ms")
        
        return metrics
    
    def evaluate_all(self, gold_data: List[Dict]) -> Dict[str, List[ModelMetrics]]:
        """Đánh giá tất cả models"""
        
        all_results = {
            "base": [],
            "reranked": [],
        }
        
        for model_key in self.config.models:
            # Kiểm tra model có tồn tại không
            if model_key not in MODEL_MAP:
                print(f"⚠️ Model không hỗ trợ: {model_key}")
                continue
            
            # Kiểm tra data có tồn tại không
            model_config = MODEL_MAP[model_key]
            vec_dir = os.path.join(self.config.vec_root, model_config["dir"])
            if not os.path.isdir(vec_dir):
                print(f"⚠️ Không tìm thấy data cho {model_key}: {vec_dir}")
                continue
            
            # Evaluate base
            try:
                base_metrics = self.evaluate_model(model_key, gold_data, use_rerank=False)
                all_results["base"].append(base_metrics)
            except Exception as e:
                print(f"❌ Lỗi evaluate {model_key}: {e}")
                continue
            
            # Evaluate với rerank
            if self.config.use_rerank and self.reranker:
                try:
                    rerank_metrics = self.evaluate_model(model_key, gold_data, use_rerank=True)
                    all_results["reranked"].append(rerank_metrics)
                except Exception as e:
                    print(f"❌ Lỗi evaluate {model_key} + rerank: {e}")
        
        return all_results


# =============================================================================
# EXCEL EXPORT
# =============================================================================
def export_to_excel(
    results: Dict[str, List[ModelMetrics]],
    output_path: str,
    config: EvalConfig,
    detailed_logs: Dict[str, List[Dict]] = None,
):
    """Xuất kết quả ra Excel"""
    
    if not HAS_PANDAS_OPENPYXL:
        print("⚠️ Cần cài pandas và openpyxl để xuất Excel")
        print("   pip install pandas openpyxl")
        # Fallback: xuất JSON
        json_path = output_path.replace(".xlsx", ".json")
        export_data = {
            "config": {
                "models": config.models,
                "topk": config.topk,
                "match_mode": config.match_mode,
                "use_rerank": config.use_rerank,
                "rerank_model": config.rerank_model,
            },
            "base_results": [
                {
                    "model": m.model_name,
                    "n_queries": m.n_queries,
                    **{f"recall@{k}": v for k, v in m.recall_at.items()},
                    "mrr": m.mrr,
                    "avg_time_ms": m.avg_time_per_query * 1000,
                }
                for m in results.get("base", [])
            ],
            "reranked_results": [
                {
                    "model": m.model_name,
                    "n_queries": m.n_queries,
                    **{f"recall@{k}": v for k, v in m.recall_at.items()},
                    "mrr": m.mrr,
                    "avg_time_ms": m.avg_time_per_query * 1000,
                }
                for m in results.get("reranked", [])
            ],
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        print(f"✓ Đã xuất kết quả ra: {json_path}")
        return
    
    # Chuẩn bị data cho DataFrame
    rows = []
    
    # Base results
    for m in results.get("base", []):
        row = {
            "Model": m.model_name,
            "Reranked": "No",
            "N": m.n_queries,
        }
        for k, v in m.recall_at.items():
            row[f"Recall@{k}"] = round(v, 4)
        row["MRR"] = round(m.mrr, 4)
        row["Avg Time (ms)"] = round(m.avg_time_per_query * 1000, 1)
        rows.append(row)
    
    # Reranked results
    for m in results.get("reranked", []):
        row = {
            "Model": m.model_name,
            "Reranked": "Yes",
            "N": m.n_queries,
        }
        for k, v in m.recall_at.items():
            row[f"Recall@{k}"] = round(v, 4)
        row["MRR"] = round(m.mrr, 4)
        row["Avg Time (ms)"] = round(m.avg_time_per_query * 1000, 1)
        rows.append(row)
    
    # Tạo DataFrame
    df = pd.DataFrame(rows)
    
    # Tạo workbook với formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Retrieval Results"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Thêm header
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    # Format header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Format data rows
    for row_idx in range(2, len(df) + 2):
        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            
            # Highlight best values (nếu là số)
            if isinstance(cell.value, (int, float)) and cell.column_letter not in ['C', 'G']:
                # Skip N và Avg Time columns cho highlighting
                pass
    
    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Thêm sheet config
    ws_config = wb.create_sheet("Config")
    config_data = [
        ["Parameter", "Value"],
        ["Models", ", ".join(config.models)],
        ["Top-K", config.topk],
        ["Match Mode", config.match_mode],
        ["Use Rerank", str(config.use_rerank)],
        ["Rerank Model", config.rerank_model],
        ["Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    for row in config_data:
        ws_config.append(row)
    
    # Format config header
    for cell in ws_config[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Thêm sheet chi tiết cho từng model
    if detailed_logs:
        # Styles cho detailed sheets
        hit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Xanh nhạt
        miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Đỏ nhạt
        
        for model_key, logs in detailed_logs.items():
            if not logs:
                continue
            
            # Tên sheet (giới hạn 31 ký tự)
            sheet_name = f"Detail_{model_key}"[:31]
            ws_detail = wb.create_sheet(sheet_name)
            
            # Columns to export
            columns = [
                "qid", "question", "gt_dieu", "gt_khoan", 
                "found_rank", "is_hit@1", "is_hit@5", "is_hit@10",
                "top1_dieu", "top1_khoan", "top1_score", "top1_text",
                "top2_dieu", "top2_khoan", "top2_score",
                "top3_dieu", "top3_khoan", "top3_score",
                "time_ms"
            ]
            
            # Header
            ws_detail.append(columns)
            for cell in ws_detail[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, log in enumerate(logs, start=2):
                row_data = [log.get(col, "") for col in columns]
                ws_detail.append(row_data)
                
                # Highlight hit/miss
                found_rank = log.get("found_rank", 0)
                if found_rank > 0 and found_rank <= 10:
                    # Hit - xanh
                    for cell in ws_detail[row_idx]:
                        cell.fill = hit_fill
                elif found_rank == 0:
                    # Miss - đỏ nhạt
                    for cell in ws_detail[row_idx]:
                        cell.fill = miss_fill
            
            # Auto-adjust column width
            for column in ws_detail.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        if cell_len > max_length:
                            max_length = cell_len
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws_detail.column_dimensions[column_letter].width = adjusted_width
        
        print(f"   Đã thêm {len(detailed_logs)} sheet chi tiết")
    
    # Lưu file
    wb.save(output_path)
    print(f"\n✓ Đã xuất kết quả ra: {output_path}")
    
    # In summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(df.to_string(index=False))


# =============================================================================
# DETAILED RESULTS EXPORT
# =============================================================================
def export_detailed_results(
    model_key: str,
    gold_data: List[Dict],
    results_per_query: List[Dict],
    output_path: str,
):
    """Xuất chi tiết kết quả từng query ra file riêng"""
    
    detailed = []
    for i, (g, res) in enumerate(zip(gold_data, results_per_query)):
        entry = {
            "qid": g.get("qid", f"Q{i+1}"),
            "question": g.get("question", ""),
            "ground_truth": g.get("ground_truth", g),
            "found_rank": res.get("found_rank", 0),
            "top_results": res.get("results", [])[:5],  # Top 5
        }
        detailed.append(entry)
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(detailed, f, ensure_ascii=False, indent=2)
    
    print(f"✓ Chi tiết: {output_path}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    parser = argparse.ArgumentParser(
        description="Đánh giá retrieval cho nhiều models với reranking",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Đánh giá tất cả models
  python eval_all_models.py --gold retrieval/gold_200_diverse.jsonl

  # Chỉ đánh giá một số models
  python eval_all_models.py --gold retrieval/goldset.jsonl --models legalhf tfidf

  # Không dùng rerank
  python eval_all_models.py --gold retrieval/gold_200_diverse.jsonl --no-rerank

  # Custom output
  python eval_all_models.py --gold retrieval/gold_200_diverse.jsonl --output my_results.xlsx
        """
    )
    
    parser.add_argument("--gold", required=True, help="Path to gold JSONL file")
    parser.add_argument("--chunks", default="output_nghidinh/chunks_clean.json",
                        help="Path to chunks JSON")
    parser.add_argument("--vec_root", default="vector_data", help="Vector data root dir")
    parser.add_argument("--models", nargs="+", default=["tfidf", "legalhf", "phobert", "dek21"],
                        choices=["tfidf", "legalhf", "phobert", "dek21"],
                        help="Models to evaluate")
    parser.add_argument("--topk", type=int, default=100, help="Top-K for retrieval (before rerank)")
    parser.add_argument("--match_mode", choices=["dieu", "dieu_khoan"], default="dieu_khoan",
                        help="Matching mode for ground truth")
    parser.add_argument("--no-rerank", action="store_true", help="Disable reranking")
    parser.add_argument("--rerank_model", default="BAAI/bge-reranker-v2-m3",
                        help="Cross-encoder model for reranking (v2-m3 tốt cho tiếng Việt)")
    parser.add_argument("--rerank_topk", type=int, default=10,
                        help="Top-K after reranking")
    parser.add_argument("--device", default="cuda", help="Device (cuda/cpu)")
    parser.add_argument("--output", default="eval_results.xlsx",
                        help="Output Excel file path")
    parser.add_argument("--detailed", action="store_true", default=True,
                        help="Ghi log chi tiết từng query vào Excel")
    parser.add_argument("--no-detailed", action="store_true",
                        help="Không ghi log chi tiết")
    
    args = parser.parse_args()
    
    # Validate inputs
    if not os.path.isfile(args.gold):
        print(f"❌ Không tìm thấy gold file: {args.gold}")
        return 1
    
    if not os.path.isfile(args.chunks):
        print(f"❌ Không tìm thấy chunks file: {args.chunks}")
        return 1
    
    # Load gold data
    print(f"Loading gold data: {args.gold}")
    gold_data = load_jsonl(args.gold)
    print(f"   Loaded {len(gold_data)} queries")
    
    # Config
    config = EvalConfig(
        models=args.models,
        topk=args.topk,
        match_mode=args.match_mode,
        use_rerank=not args.no_rerank,
        rerank_model=args.rerank_model,
        rerank_topk=args.rerank_topk,
        device=args.device,
        vec_root=args.vec_root,
        chunks_path=args.chunks,
        save_detailed=not args.no_detailed,
    )
    
    print(f"\nConfig:")
    print(f"   Models: {config.models}")
    print(f"   Top-K retrieve: {config.topk}")
    print(f"   Rerank top-K: {config.rerank_topk}")
    print(f"   Match mode: {config.match_mode}")
    print(f"   Use rerank: {config.use_rerank}")
    print(f"   Save detailed: {config.save_detailed}")
    if config.use_rerank:
        print(f"   Rerank model: {config.rerank_model}")
    
    # Evaluate
    evaluator = RetrievalEvaluator(config)
    results = evaluator.evaluate_all(gold_data)
    
    # Export với detailed logs
    detailed_logs = evaluator.detailed_logs if config.save_detailed else None
    export_to_excel(results, args.output, config, detailed_logs)
    
    return 0


if __name__ == "__main__":
    exit(main())
