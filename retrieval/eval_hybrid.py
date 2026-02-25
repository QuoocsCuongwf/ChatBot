"""
Hybrid Retrieval: Dense + TF-IDF
=================================
Kết hợp semantic search (dense) với keyword matching (TF-IDF)
để tận dụng ưu điểm của cả hai:
- Dense: Hiểu ngữ nghĩa, paraphrase
- TF-IDF: Match chính xác thuật ngữ pháp lý

Formula: score = alpha * dense_score + (1-alpha) * tfidf_score

Usage:
  python eval_hybrid.py --gold retrieval/gold_200_diverse.jsonl
  python eval_hybrid.py --gold retrieval/gold_200_diverse.jsonl --alpha 0.6
  python eval_hybrid.py --gold retrieval/gold_200_diverse.jsonl --append_to eval_full_detailed.xlsx
"""

import os
import json
import argparse
import time
from datetime import datetime
from typing import List, Dict, Any, Tuple
from dataclasses import dataclass, field
from collections import defaultdict

import numpy as np

# Dense retrieval
import faiss
from sentence_transformers import SentenceTransformer

# TF-IDF
import joblib
import scipy.sparse as sp
from sklearn.metrics.pairwise import linear_kernel

# Reranking
try:
    from sentence_transformers import CrossEncoder
    HAS_CROSS_ENCODER = True
except ImportError:
    HAS_CROSS_ENCODER = False

# Excel
try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# CONFIG
# =============================================================================
@dataclass
class HybridConfig:
    """Cấu hình hybrid retrieval"""
    dense_model: str = "legalhf"
    alpha: float = 0.5  # weight cho dense, (1-alpha) cho tfidf
    topk_retrieve: int = 100  # retrieve wide từ mỗi model
    topk_final: int = 50  # sau khi merge
    use_rerank: bool = True
    rerank_model: str = "BAAI/bge-reranker-base"
    rerank_topk: int = 10
    match_mode: str = "dieu_khoan"
    device: str = "cuda"
    vec_root: str = "vector_data"
    chunks_path: str = "output_nghidinh/chunks_clean_norm.json"


# Model paths
DENSE_CONFIG = {
    "legalhf": {
        "dir": "legal_hf_cosine",
        "encoder": "Quockhanh05/Vietnam_legal_embeddings",
    },
    "phobert": {
        "dir": "phobert",
        "encoder": "vinai/phobert-base",
    },
    "dek21": {
        "dir": "dek21",
        "encoder": "dangvantuan/vietnamese-embedding",
    },
}

TFIDF_DIR = "tfidf"


# =============================================================================
# HELPERS
# =============================================================================
def load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def load_jsonl(path: str) -> List[Dict]:
    items = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                items.append(json.loads(line))
    return items

def norm_str(x) -> str:
    if x is None:
        return ""
    return str(x).strip().lower()

def l2_normalize(x: np.ndarray, eps: float = 1e-12) -> np.ndarray:
    n = np.linalg.norm(x, axis=1, keepdims=True)
    return x / np.clip(n, eps, None)

def match_ground_truth(gt: Dict, meta: Dict, mode: str = "dieu_khoan") -> bool:
    """So khớp ground truth với metadata của chunk"""
    gt_van_ban = gt.get("van_ban", "") or gt.get("source_contains", "")
    meta_van_ban = meta.get("van_ban", "")
    
    gt_vb_norm = norm_str(gt_van_ban)
    meta_vb_norm = norm_str(meta_van_ban)
    
    vb_match = False
    if gt_vb_norm and meta_vb_norm:
        if gt_vb_norm == meta_vb_norm:
            vb_match = True
        elif gt_vb_norm in meta_vb_norm or meta_vb_norm in gt_vb_norm:
            vb_match = True
    elif not gt_vb_norm:
        vb_match = True
    
    if not vb_match:
        return False
    
    if norm_str(gt.get("dieu")) != norm_str(meta.get("dieu")):
        return False
    
    if mode == "dieu":
        return True
    
    gt_khoan = gt.get("khoan")
    meta_khoan = meta.get("khoan")
    
    if gt_khoan is None:
        return True
    
    return norm_str(gt_khoan) == norm_str(meta_khoan)

def mrr_from_ranks(ranks: List[int]) -> float:
    if not ranks:
        return 0.0
    s = sum(1.0/r if r > 0 else 0.0 for r in ranks)
    return s / len(ranks)


# =============================================================================
# SCORE NORMALIZATION
# =============================================================================
def min_max_normalize(scores: np.ndarray) -> np.ndarray:
    """Normalize scores to [0, 1]"""
    min_s = scores.min()
    max_s = scores.max()
    if max_s - min_s < 1e-9:
        return np.ones_like(scores) * 0.5
    return (scores - min_s) / (max_s - min_s)

def z_score_normalize(scores: np.ndarray) -> np.ndarray:
    """Z-score normalization"""
    mean = scores.mean()
    std = scores.std()
    if std < 1e-9:
        return np.zeros_like(scores)
    return (scores - mean) / std


# =============================================================================
# HYBRID RETRIEVER
# =============================================================================
class HybridRetriever:
    """Kết hợp Dense + TF-IDF retrieval"""
    
    def __init__(self, config: HybridConfig):
        self.config = config
        self.chunks = load_json(config.chunks_path)
        
        # Load components
        self._load_dense()
        self._load_tfidf()
        
        # Reranker
        self.reranker = None
        if config.use_rerank and HAS_CROSS_ENCODER:
            try:
                print(f"Loading reranker: {config.rerank_model}...")
                self.reranker = CrossEncoder(config.rerank_model, device=config.device)
            except Exception as e:
                print(f"⚠️ Không load được reranker: {e}")
    
    def _load_dense(self):
        """Load dense model và index"""
        dense_cfg = DENSE_CONFIG[self.config.dense_model]
        vec_dir = os.path.join(self.config.vec_root, dense_cfg["dir"])
        
        print(f"Loading dense encoder: {dense_cfg['encoder']}...")
        self.dense_encoder = SentenceTransformer(dense_cfg["encoder"], device=self.config.device)
        
        idx_path = os.path.join(vec_dir, "index.faiss")
        print(f"Loading FAISS index: {idx_path}...")
        self.dense_index = faiss.read_index(idx_path)
    
    def _load_tfidf(self):
        """Load TF-IDF vectorizer và matrix"""
        vec_dir = os.path.join(self.config.vec_root, TFIDF_DIR)
        
        vec_path = os.path.join(vec_dir, "vectorizer.pkl")
        if not os.path.isfile(vec_path):
            vec_path = os.path.join(vec_dir, "vectorizer.joblib")
        
        mat_path = os.path.join(vec_dir, "tfidf_matrix.npz")
        
        print(f"Loading TF-IDF: {vec_dir}...")
        self.tfidf_vectorizer = joblib.load(vec_path)
        self.tfidf_matrix = sp.load_npz(mat_path)
    
    def _dense_retrieve(self, query: str, topk: int) -> Dict[int, float]:
        """Retrieve với dense model, trả về {chunk_idx: score}"""
        q_vec = self.dense_encoder.encode([query], convert_to_numpy=True).astype("float32")
        q_vec = l2_normalize(q_vec)
        
        scores, ids = self.dense_index.search(q_vec, topk)
        scores = scores[0]
        ids = ids[0]
        
        result = {}
        for idx, sc in zip(ids, scores):
            if idx >= 0 and idx < len(self.chunks):
                result[int(idx)] = float(sc)
        return result
    
    def _tfidf_retrieve(self, query: str, topk: int) -> Dict[int, float]:
        """Retrieve với TF-IDF, trả về {chunk_idx: score}"""
        q_vec = self.tfidf_vectorizer.transform([query])
        sims = linear_kernel(q_vec, self.tfidf_matrix).ravel()
        
        if topk >= len(sims):
            top_idx = np.argsort(-sims)
        else:
            top_idx = np.argpartition(-sims, topk)[:topk]
            top_idx = top_idx[np.argsort(-sims[top_idx])]
        
        result = {}
        for idx in top_idx[:topk]:
            result[int(idx)] = float(sims[int(idx)])
        return result
    
    def hybrid_retrieve(
        self,
        query: str,
        topk: int = 50,
        alpha: float = 0.5,
        normalize_method: str = "minmax",
    ) -> List[Dict]:
        """
        Hybrid retrieval kết hợp dense + tfidf
        
        Args:
            query: Query string
            topk: Số kết quả trả về
            alpha: Weight cho dense (0-1), tfidf = 1-alpha
            normalize_method: "minmax" hoặc "zscore"
        
        Returns:
            List of results sorted by hybrid score
        """
        # Retrieve từ cả 2 sources
        dense_scores = self._dense_retrieve(query, self.config.topk_retrieve)
        tfidf_scores = self._tfidf_retrieve(query, self.config.topk_retrieve)
        
        # Collect all unique chunk indices
        all_indices = set(dense_scores.keys()) | set(tfidf_scores.keys())
        
        if not all_indices:
            return []
        
        # Convert to arrays for normalization
        indices = list(all_indices)
        dense_arr = np.array([dense_scores.get(i, 0.0) for i in indices])
        tfidf_arr = np.array([tfidf_scores.get(i, 0.0) for i in indices])
        
        # Normalize scores
        if normalize_method == "minmax":
            dense_norm = min_max_normalize(dense_arr)
            tfidf_norm = min_max_normalize(tfidf_arr)
        else:
            dense_norm = z_score_normalize(dense_arr)
            tfidf_norm = z_score_normalize(tfidf_arr)
        
        # Combine scores
        hybrid_scores = alpha * dense_norm + (1 - alpha) * tfidf_norm
        
        # Sort by hybrid score
        sorted_indices = np.argsort(-hybrid_scores)[:topk]
        
        results = []
        for rank, idx in enumerate(sorted_indices, start=1):
            chunk_idx = indices[idx]
            chunk = self.chunks[chunk_idx]
            
            results.append({
                "rank": rank,
                "score": float(hybrid_scores[idx]),
                "dense_score": float(dense_arr[idx]),
                "dense_norm": float(dense_norm[idx]),
                "tfidf_score": float(tfidf_arr[idx]),
                "tfidf_norm": float(tfidf_norm[idx]),
                "text": chunk.get("text", ""),
                "metadata": chunk.get("metadata", {}),
                "chunk_idx": chunk_idx,
            })
        
        return results
    
    def rerank(self, query: str, candidates: List[Dict], topk: int = 10) -> List[Dict]:
        """Rerank candidates với cross-encoder"""
        if not self.reranker or not candidates:
            return candidates[:topk]
        
        pairs = [(query, c.get("text", "")) for c in candidates]
        scores = self.reranker.predict(pairs)
        
        scored = []
        for i, c in enumerate(candidates):
            scored.append({
                **c,
                "original_score": c["score"],
                "rerank_score": float(scores[i]),
            })
        
        scored.sort(key=lambda x: x["rerank_score"], reverse=True)
        
        results = []
        for new_rank, item in enumerate(scored[:topk], start=1):
            item["rank"] = new_rank
            item["score"] = item["rerank_score"]
            results.append(item)
        
        return results


# =============================================================================
# EVALUATOR
# =============================================================================
class HybridEvaluator:
    """Đánh giá hybrid retrieval"""
    
    def __init__(self, retriever: HybridRetriever, config: HybridConfig):
        self.retriever = retriever
        self.config = config
        self.detailed_logs = {}
    
    def evaluate(
        self,
        gold_data: List[Dict],
        alpha: float = 0.5,
        use_rerank: bool = False,
    ) -> Tuple[Dict, List[Dict]]:
        """
        Đánh giá hybrid retrieval
        
        Returns:
            (metrics_dict, query_details_list)
        """
        name = f"hybrid_a{alpha}" + ("_rerank" if use_rerank else "")
        print(f"\n{'='*60}")
        print(f"Evaluating: Hybrid (alpha={alpha}) {'+ Rerank' if use_rerank else ''}")
        print(f"{'='*60}")
        
        Ks = [1, 3, 5, 10]
        hit_at = {k: 0 for k in Ks}
        ranks = []
        total_time = 0.0
        query_details = []
        
        for i, g in enumerate(gold_data):
            query = g.get("question", "")
            qid = g.get("qid", f"Q{i+1}")
            
            gt_list = g.get("ground_truth", [g])
            if not isinstance(gt_list, list):
                gt_list = [gt_list]
            
            gt_dieu = gt_list[0].get("dieu", "") if gt_list else ""
            gt_khoan = gt_list[0].get("khoan", "") if gt_list else ""
            gt_van_ban = gt_list[0].get("van_ban", "") if gt_list else ""
            
            start_time = time.time()
            
            # Hybrid retrieve
            results = self.retriever.hybrid_retrieve(
                query,
                topk=self.config.topk_final,
                alpha=alpha,
            )
            
            # Rerank
            if use_rerank and self.retriever.reranker:
                results = self.retriever.rerank(query, results, topk=self.config.rerank_topk)
            
            elapsed = time.time() - start_time
            total_time += elapsed
            
            # Find first correct
            found_rank = 0
            for r in results:
                meta = r.get("metadata", {})
                for gt in gt_list:
                    if match_ground_truth(gt, meta, self.config.match_mode):
                        found_rank = r["rank"]
                        break
                if found_rank > 0:
                    break
            
            ranks.append(found_rank)
            
            for k in Ks:
                if 0 < found_rank <= k:
                    hit_at[k] += 1
            
            # Log detail
            top1 = results[0] if results else {}
            top1_meta = top1.get("metadata", {})
            detail = {
                "qid": qid,
                "question": query[:100],
                "gt_dieu": gt_dieu,
                "gt_khoan": gt_khoan,
                "gt_van_ban": gt_van_ban[:50] if gt_van_ban else "",
                "found_rank": found_rank,
                "is_hit@1": found_rank == 1,
                "is_hit@5": 0 < found_rank <= 5,
                "is_hit@10": 0 < found_rank <= 10,
                "top1_dieu": top1_meta.get("dieu", ""),
                "top1_khoan": top1_meta.get("khoan", ""),
                "top1_score": round(top1.get("score", 0), 4),
                "top1_dense": round(top1.get("dense_norm", 0), 4),
                "top1_tfidf": round(top1.get("tfidf_norm", 0), 4),
                "top1_text": top1.get("text", "")[:100],
                "time_ms": round(elapsed * 1000, 1),
            }
            
            # Top 3
            for j in range(min(3, len(results))):
                r = results[j]
                m = r.get("metadata", {})
                detail[f"top{j+1}_dieu"] = m.get("dieu", "")
                detail[f"top{j+1}_khoan"] = m.get("khoan", "")
                detail[f"top{j+1}_score"] = round(r.get("score", 0), 4)
            
            query_details.append(detail)
            
            if (i + 1) % 50 == 0 or (i + 1) == len(gold_data):
                print(f"   Processed: {i+1}/{len(gold_data)}")
        
        # Metrics
        n = len(gold_data)
        recall_at = {k: hit_at[k] / n if n > 0 else 0.0 for k in Ks}
        mrr = mrr_from_ranks(ranks)
        avg_time = total_time / n if n > 0 else 0.0
        
        metrics = {
            "model": f"hybrid_a{alpha}",
            "reranked": use_rerank,
            "alpha": alpha,
            "n_queries": n,
            "recall_at": recall_at,
            "mrr": mrr,
            "hit_at": hit_at,
            "avg_time_ms": avg_time * 1000,
        }
        
        print(f"\n   Results for Hybrid (alpha={alpha}) {'+ Rerank' if use_rerank else ''}:")
        for k in Ks:
            print(f"      Recall@{k}: {recall_at[k]:.4f} ({hit_at[k]}/{n})")
        print(f"      MRR: {mrr:.4f}")
        print(f"      Avg time/query: {avg_time*1000:.1f}ms")
        
        self.detailed_logs[name] = query_details
        
        return metrics, query_details


# =============================================================================
# EXCEL EXPORT
# =============================================================================
def append_to_excel(
    output_path: str,
    metrics_list: List[Dict],
    detailed_logs: Dict[str, List[Dict]],
):
    """Thêm kết quả hybrid vào file Excel đã có"""
    
    if not HAS_OPENPYXL:
        print("⚠️ Cần openpyxl để ghi Excel")
        return
    
    # Load existing workbook hoặc tạo mới
    if os.path.isfile(output_path):
        print(f"Appending to existing: {output_path}")
        wb = load_workbook(output_path)
    else:
        print(f"Creating new: {output_path}")
        wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    hit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    hybrid_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Vàng nhạt
    
    # Update/Add Results sheet
    if "Retrieval Results" in wb.sheetnames:
        ws = wb["Retrieval Results"]
        # Find last row
        last_row = ws.max_row
    else:
        ws = wb.active
        ws.title = "Retrieval Results"
        # Add header
        headers = ["Model", "Reranked", "Alpha", "N", "Recall@1", "Recall@3", "Recall@5", "Recall@10", "MRR", "Avg Time (ms)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        last_row = 1
    
    # Add metrics
    for m in metrics_list:
        row = [
            m["model"],
            "Yes" if m["reranked"] else "No",
            m.get("alpha", ""),
            m["n_queries"],
            round(m["recall_at"].get(1, 0), 4),
            round(m["recall_at"].get(3, 0), 4),
            round(m["recall_at"].get(5, 0), 4),
            round(m["recall_at"].get(10, 0), 4),
            round(m["mrr"], 4),
            round(m["avg_time_ms"], 1),
        ]
        ws.append(row)
        # Highlight hybrid rows
        for cell in ws[ws.max_row]:
            cell.fill = hybrid_fill
    
    # Add detailed sheets
    for log_name, logs in detailed_logs.items():
        if not logs:
            continue
        
        sheet_name = f"Detail_{log_name}"[:31]
        
        # Remove if exists
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        ws_detail = wb.create_sheet(sheet_name)
        
        columns = [
            "qid", "question", "gt_dieu", "gt_khoan",
            "found_rank", "is_hit@1", "is_hit@5", "is_hit@10",
            "top1_dieu", "top1_khoan", "top1_score", "top1_dense", "top1_tfidf",
            "top1_text", "time_ms",
        ]
        
        ws_detail.append(columns)
        for cell in ws_detail[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, log in enumerate(logs, start=2):
            row_data = [log.get(col, "") for col in columns]
            ws_detail.append(row_data)
            
            found_rank = log.get("found_rank", 0)
            if 0 < found_rank <= 10:
                for cell in ws_detail[row_idx]:
                    cell.fill = hit_fill
            elif found_rank == 0:
                for cell in ws_detail[row_idx]:
                    cell.fill = miss_fill
        
        # Auto width
        for column in ws_detail.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws_detail.column_dimensions[col_letter].width = min(max_len + 2, 40)
    
    wb.save(output_path)
    print(f"\n✓ Đã lưu kết quả vào: {output_path}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    parser = argparse.ArgumentParser(
        description="Hybrid Retrieval: Dense + TF-IDF",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Đánh giá với alpha=0.5 (mặc định)
  python eval_hybrid.py --gold retrieval/gold_200_diverse.jsonl
  
  # Thử nhiều giá trị alpha
  python eval_hybrid.py --gold retrieval/gold_200_diverse.jsonl --alphas 0.3 0.5 0.7
  
  # Append kết quả vào file Excel đã có
  python eval_hybrid.py --gold retrieval/gold_200_diverse.jsonl --append_to eval_full_detailed.xlsx
        """
    )
    
    parser.add_argument("--gold", required=True, help="Gold JSONL file")
    parser.add_argument("--chunks", default="output_nghidinh/chunks_clean_norm.json")
    parser.add_argument("--vec_root", default="vector_data")
    parser.add_argument("--dense_model", default="legalhf", choices=["legalhf", "phobert", "dek21"])
    parser.add_argument("--alphas", nargs="+", type=float, default=[0.5],
                        help="Alpha values to try (default: 0.5)")
    parser.add_argument("--topk", type=int, default=100, help="Top-K retrieve từ mỗi source")
    parser.add_argument("--topk_final", type=int, default=50, help="Top-K sau merge")
    parser.add_argument("--no-rerank", action="store_true")
    parser.add_argument("--rerank_model", default="BAAI/bge-reranker-base")
    parser.add_argument("--rerank_topk", type=int, default=10)
    parser.add_argument("--device", default="cuda")
    parser.add_argument("--output", default="eval_hybrid.xlsx")
    parser.add_argument("--append_to", default=None,
                        help="Append results to existing Excel file")
    
    args = parser.parse_args()
    
    # Validate
    if not os.path.isfile(args.gold):
        print(f"❌ Không tìm thấy: {args.gold}")
        return 1
    
    if not os.path.isfile(args.chunks):
        print(f"❌ Không tìm thấy: {args.chunks}")
        return 1
    
    # Load gold
    print(f"Loading gold: {args.gold}")
    gold_data = load_jsonl(args.gold)
    print(f"   Loaded {len(gold_data)} queries")
    
    # Config
    config = HybridConfig(
        dense_model=args.dense_model,
        topk_retrieve=args.topk,
        topk_final=args.topk_final,
        use_rerank=not args.no_rerank,
        rerank_model=args.rerank_model,
        rerank_topk=args.rerank_topk,
        device=args.device,
        vec_root=args.vec_root,
        chunks_path=args.chunks,
    )
    
    print(f"\nConfig:")
    print(f"   Dense model: {config.dense_model}")
    print(f"   Alphas to test: {args.alphas}")
    print(f"   Top-K retrieve: {config.topk_retrieve}")
    print(f"   Top-K final: {config.topk_final}")
    print(f"   Use rerank: {config.use_rerank}")
    
    # Initialize
    retriever = HybridRetriever(config)
    evaluator = HybridEvaluator(retriever, config)
    
    # Evaluate with different alphas
    all_metrics = []
    all_logs = {}
    
    for alpha in args.alphas:
        # Base hybrid
        metrics, _ = evaluator.evaluate(gold_data, alpha=alpha, use_rerank=False)
        all_metrics.append(metrics)
        
        # With rerank
        if config.use_rerank:
            metrics_rerank, _ = evaluator.evaluate(gold_data, alpha=alpha, use_rerank=True)
            all_metrics.append(metrics_rerank)
    
    all_logs = evaluator.detailed_logs
    
    # Export
    output_file = args.append_to or args.output
    append_to_excel(output_file, all_metrics, all_logs)
    
    # Summary
    print("\n" + "="*70)
    print("HYBRID RETRIEVAL SUMMARY")
    print("="*70)
    
    for m in all_metrics:
        rerank_str = "+ Rerank" if m["reranked"] else ""
        print(f"\n{m['model']} {rerank_str}:")
        print(f"   Recall@1: {m['recall_at'][1]:.4f}")
        print(f"   Recall@5: {m['recall_at'][5]:.4f}")
        print(f"   Recall@10: {m['recall_at'][10]:.4f}")
        print(f"   MRR: {m['mrr']:.4f}")
    
    return 0


if __name__ == "__main__":
    exit(main())
