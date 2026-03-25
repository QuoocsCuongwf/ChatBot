"""
pipeline_logger.py — Nhật ký chuyên nghiệp cho Legal RAG Pipeline

Ghi log toàn bộ pipeline vào Excel (.xlsx) + CSV backup.
Mỗi query = 1 dòng, ghi đầy đủ:
  - Timestamp, Query
  - Tier (LOCAL / API / NONE)
  - Gating decision, scores, margin
  - Retrieved chunks, reranked scores
  - LLM response, latency
  - Citations, errors

Usage:
    logger = PipelineLogger(log_dir="cross-encoder/generation/outputs/logs")
    logger.log(entry)
    logger.flush()   # ghi xuống file
"""

import csv
import json
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Any
from dataclasses import dataclass, field, asdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════════════
# LOG ENTRY
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class PipelineLogEntry:
    """Một dòng log cho pipeline."""

    # Identifiers
    timestamp: str = ""
    run_date: str = ""                  # Ngày giờ lần chạy (phân biệt các lần chạy)
    query_id: str = ""
    query: str = ""

    # Retrieval
    num_chunks_retrieved: int = 0
    num_chunks_after_dedup: int = 0
    top1_chunk_id: int = -1
    top1_text_preview: str = ""         # 100 ký tự đầu

    # Scores
    score_retrieval_top1: float = 0.0
    score_rerank_top1: float = 0.0
    score_rerank_top2: float = 0.0
    margin: float = 0.0

    # Gating
    gating_decision: str = ""           # answer / cautious / abstain / ask_back
    gating_reason: str = ""
    gating_confidence: float = 0.0

    # 2-Tier routing
    tier: str = ""                      # local / api / none
    tier_reason: str = ""               # Lý do chọn tier
    # Enhanced metrics
    confidence_final: float = 0.0       # sigmoid(margin) — normalized confidence
    semantic_similarity: float = 0.0    # sigmoid(score/5) — normalized semantic sim
    query_type: str = ""                # detected query type (định nghĩa, thủ tục, ...)
    context_token_length: int = 0       # ước tính token length of context
    lexical_overlap: float = 0.0        # lexical overlap query↔context
    # LLM Generation
    llm_backend: str = ""               # gemini / llama_cpp / placeholder
    llm_model: str = ""
    generated_answer: str = ""
    answer_length: int = 0

    # Citations
    num_citations: int = 0
    citations_str: str = ""             # "Điều 5 Khoản 2, Điều 8 Khoản 1"
    citation_hit: bool = False          # Trúng expected citation?

    # Latency (ms)
    latency_retrieve_ms: float = 0.0
    latency_rerank_ms: float = 0.0
    latency_context_ms: float = 0.0
    latency_llm_ms: float = 0.0
    latency_total_ms: float = 0.0

    # Anti-429 tracking
    api_failed: bool = False            # API gọi thất bại (429/5xx) sau retries?
    fallback_to_local: bool = False     # Đã fallback sang local?
    retry_count: int = 0                # Số lần retry

    # Status
    error: str = ""
    status: str = "ok"                  # ok / error / abstain / ask_back

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def to_flat_row(self) -> List[Any]:
        """Trả về list giá trị theo thứ tự cột."""
        d = self.to_dict()
        return [d[col] for col in COLUMN_ORDER]


# Thứ tự cột cho Excel/CSV
COLUMN_ORDER = [
    "timestamp", "run_date", "query_id", "query",
    "tier", "tier_reason",
    "gating_decision", "gating_reason", "gating_confidence",
    "confidence_final", "semantic_similarity", "lexical_overlap",
    "query_type", "context_token_length",
    "score_rerank_top1", "score_rerank_top2", "margin",
    "score_retrieval_top1",
    "num_chunks_retrieved", "num_chunks_after_dedup",
    "top1_chunk_id", "top1_text_preview",
    "llm_backend", "llm_model",
    "generated_answer", "answer_length",
    "num_citations", "citations_str", "citation_hit",
    "latency_retrieve_ms", "latency_rerank_ms",
    "latency_context_ms", "latency_llm_ms", "latency_total_ms",
    "api_failed", "fallback_to_local", "retry_count",
    "error", "status",
]

# Headers tiếng Việt đẹp cho Excel
COLUMN_HEADERS_VI = {
    "timestamp":              "Thời gian",
    "run_date":               "Ngày chạy",
    "query_id":               "Mã câu hỏi",
    "query":                  "Câu hỏi",
    "tier":                   "Tầng LLM",
    "tier_reason":            "Lý do chọn tầng",
    "gating_decision":        "Quyết định Gating",
    "gating_reason":          "Chi tiết Gating",
    "gating_confidence":      "Confidence",
    "confidence_final":        "Confidence Final (sigmoid)",
    "semantic_similarity":     "Semantic Similarity",
    "lexical_overlap":         "Lexical Overlap (%)",
    "query_type":              "Loại câu hỏi",
    "context_token_length":    "Context Tokens (est.)",
    "score_rerank_top1":      "Score Top-1 (CE)",
    "score_rerank_top2":      "Score Top-2 (CE)",
    "margin":                 "Margin (Top1-Top2)",
    "score_retrieval_top1":   "Score Retrieval",
    "num_chunks_retrieved":   "Chunks retrieved",
    "num_chunks_after_dedup": "Chunks sau dedup",
    "top1_chunk_id":          "Chunk ID top-1",
    "top1_text_preview":      "Nội dung top-1 (preview)",
    "llm_backend":            "LLM Backend",
    "llm_model":              "LLM Model",
    "generated_answer":       "Câu trả lời",
    "answer_length":          "Độ dài trả lời",
    "num_citations":          "Số trích dẫn",
    "citations_str":          "Trích dẫn",
    "citation_hit":           "Citation Hit?",
    "latency_retrieve_ms":    "Thời gian Retrieve (ms)",
    "latency_rerank_ms":      "Thời gian Rerank (ms)",
    "latency_context_ms":     "Thời gian Context (ms)",
    "latency_llm_ms":         "Thời gian LLM (ms)",
    "latency_total_ms":       "Tổng thời gian (ms)",
    "error":                  "Lỗi",
    "status":                 "Trạng thái",
}

# Màu cho tier
TIER_COLORS = {
    "local": "C6EFCE",     # Xanh lá nhạt
    "api":   "BDD7EE",     # Xanh dương nhạt
    "none":  "F2DCDB",     # Đỏ nhạt
}

DECISION_COLORS = {
    "answer":   "C6EFCE",
    "cautious": "FFEB9C",
    "abstain":  "F2DCDB",
    "ask_back": "E4DFEC",
}


# ═══════════════════════════════════════════════════════════════════════════════
# PIPELINE LOGGER
# ═══════════════════════════════════════════════════════════════════════════════

class PipelineLogger:
    """
    Logger chuyên nghiệp cho Legal RAG Pipeline.
    
    Ghi log vào:
    1. Excel (.xlsx) — có màu, freeze panes, auto-width
    2. CSV (.csv) — backup dễ mở
    3. JSONL (.jsonl) — raw structured data
    """

    def __init__(
        self,
        log_dir: Optional[str] = None,
        run_id: Optional[str] = None,
        auto_flush_every: int = 10,
        append: bool = False,
    ):
        self.log_dir = Path(log_dir or "cross-encoder/generation/outputs/logs")
        self.log_dir.mkdir(parents=True, exist_ok=True)

        self.append = append
        self.entries: List[PipelineLogEntry] = []
        self.auto_flush_every = auto_flush_every
        self._flush_count = 0
        self._run_date = datetime.now().strftime("%Y-%m-%d %H:%M")

        if append:
            # Append mode: tên file cố định, ghi tiếp vào file cũ
            self.run_id = run_id or "pipeline_log"
            self.xlsx_path = self.log_dir / f"{self.run_id}.xlsx"
            self.csv_path = self.log_dir / f"{self.run_id}.csv"
            self.jsonl_path = self.log_dir / f"{self.run_id}.jsonl"
            # Load existing entries from JSONL (for Excel rebuild)
            self._prior_entries = self._load_existing_jsonl()
            # Chỉ tạo CSV header nếu file chưa tồn tại
            if not self.csv_path.exists():
                self._init_csv()
            print(f"[Logger] Append mode: {self.run_id} ({len(self._prior_entries)} existing entries)")
        else:
            # New-file mode: mỗi run tạo file riêng (như cũ)
            self.run_id = run_id or datetime.now().strftime("run_%Y%m%d_%H%M%S")
            self._prior_entries = []
            self.xlsx_path = self.log_dir / f"{self.run_id}.xlsx"
            self.csv_path = self.log_dir / f"{self.run_id}.csv"
            self.jsonl_path = self.log_dir / f"{self.run_id}.jsonl"
            self._init_csv()
            print(f"[Logger] Initialized: {self.run_id}")

        print(f"[Logger] Excel: {self.xlsx_path}")
        print(f"[Logger] CSV:   {self.csv_path}")

    def _init_csv(self):
        """Tạo CSV file với headers."""
        with open(self.csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            headers = [COLUMN_HEADERS_VI.get(col, col) for col in COLUMN_ORDER]
            writer.writerow(headers)

    def log(self, entry: PipelineLogEntry):
        """Thêm 1 entry vào log."""
        if not entry.timestamp:
            entry.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not entry.run_date:
            entry.run_date = self._run_date

        self.entries.append(entry)

        # Append to CSV immediately (safe)
        self._append_csv(entry)

        # Append to JSONL immediately
        self._append_jsonl(entry)

        self._flush_count += 1
        if self._flush_count >= self.auto_flush_every:
            self.flush_excel()
            self._flush_count = 0

    def _append_csv(self, entry: PipelineLogEntry):
        """Ghi 1 dòng vào CSV."""
        try:
            with open(self.csv_path, "a", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(entry.to_flat_row())
        except PermissionError:
            print(f"[Logger] Warning: Could not append to CSV (Permission Denied). Is {self.csv_path.name} open?")
        except Exception as e:
            print(f"[Logger] Warning: CSV append failed: {e}")

    def _append_jsonl(self, entry: PipelineLogEntry):
        """Ghi 1 dòng vào JSONL."""
        try:
            with open(self.jsonl_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(entry.to_dict(), ensure_ascii=False) + "\n")
        except PermissionError:
            # Silently ignore or print short warning
            pass
        except Exception as e:
            print(f"[Logger] Warning: JSONL append failed: {e}")

    def _load_existing_jsonl(self) -> List[PipelineLogEntry]:
        """Load entries từ JSONL file đã tồn tại (cho append mode)."""
        entries = []
        if self.jsonl_path.exists():
            try:
                with open(self.jsonl_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        d = json.loads(line)
                        entry = PipelineLogEntry(**{k: v for k, v in d.items()
                                                    if k in PipelineLogEntry.__dataclass_fields__})
                        entries.append(entry)
            except Exception as e:
                print(f"[Logger] Warning: could not load existing JSONL: {e}")
        return entries

    def flush_excel(self):
        """Ghi toàn bộ entries ra Excel (overwrite). Append mode gộp cả entries cũ."""
        if not HAS_OPENPYXL:
            print("[Logger] openpyxl not installed, skipping Excel. pip install openpyxl")
            return

        # Trong append mode: gộp entries cũ + entries mới cho Excel
        original_entries = self.entries
        if self.append and self._prior_entries:
            self.entries = self._prior_entries + self.entries

        wb = Workbook()

        # ─── Sheet 1: Chi tiết ───
        ws = wb.active
        ws.title = "Chi tiết Pipeline"
        self._write_detail_sheet(ws)

        # ─── Sheet 2: Thống kê ───
        ws_stats = wb.create_sheet("Thống kê tổng hợp")
        self._write_stats_sheet(ws_stats)

        # ─── Sheet 3: Tier Distribution ───
        ws_tier = wb.create_sheet("Phân bố Tier")
        self._write_tier_sheet(ws_tier)

        total = len(self.entries)
        new_count = len(original_entries) if self.append else total
        saved_path = self._save_excel_safe(wb)
        if saved_path:
            print(f"[Logger] Excel saved: {saved_path} ({total} total, {new_count} new)")
        else:
            print(f"[Logger] ⚠ Không ghi được Excel (file đang mở?). CSV/JSONL vẫn OK.")

        # Restore entries to only current run's entries
        if self.append and self._prior_entries:
            self.entries = original_entries

    def _save_excel_safe(self, wb) -> Optional[str]:
        """Ghi Excel an toàn — thử nhiều path nếu file bị lock."""
        # Thử 1: path chính
        candidates = [self.xlsx_path]
        # Thử 2-4: path với timestamp suffix
        ts = datetime.now().strftime("%H%M%S")
        candidates.append(self.xlsx_path.with_stem(f"{self.xlsx_path.stem}_{ts}"))
        candidates.append(self.xlsx_path.with_stem(f"{self.xlsx_path.stem}_new"))

        for path in candidates:
            try:
                wb.save(str(path))
                return str(path)
            except PermissionError:
                continue
        return None

    def _write_detail_sheet(self, ws):
        """Ghi sheet chi tiết."""
        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Write headers
        for col_idx, col_name in enumerate(COLUMN_ORDER, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = COLUMN_HEADERS_VI.get(col_name, col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data
        for row_idx, entry in enumerate(self.entries, 2):
            row_data = entry.to_flat_row()
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Truncate long text for Excel
                if isinstance(value, str) and len(value) > 500:
                    cell.value = value[:497] + "..."
                else:
                    cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=False)

            # Color tier column
            tier_col = COLUMN_ORDER.index("tier") + 1
            tier_val = str(entry.tier).lower()
            if tier_val in TIER_COLORS:
                ws.cell(row=row_idx, column=tier_col).fill = PatternFill(
                    start_color=TIER_COLORS[tier_val],
                    end_color=TIER_COLORS[tier_val],
                    fill_type="solid"
                )

            # Color decision column
            dec_col = COLUMN_ORDER.index("gating_decision") + 1
            dec_val = str(entry.gating_decision).lower()
            if dec_val in DECISION_COLORS:
                ws.cell(row=row_idx, column=dec_col).fill = PatternFill(
                    start_color=DECISION_COLORS[dec_val],
                    end_color=DECISION_COLORS[dec_val],
                    fill_type="solid"
                )

        # Auto-fit column widths (approx)
        col_widths = {
            "timestamp": 20, "run_date": 18, "query_id": 10, "query": 50,
            "tier": 8, "tier_reason": 35,
            "gating_decision": 12, "gating_reason": 40, "gating_confidence": 12,
            "score_rerank_top1": 14, "score_rerank_top2": 14, "margin": 14,
            "score_retrieval_top1": 14,
            "num_chunks_retrieved": 10, "num_chunks_after_dedup": 10,
            "top1_chunk_id": 10, "top1_text_preview": 60,
            "llm_backend": 12, "llm_model": 18,
            "generated_answer": 60, "answer_length": 10,
            "num_citations": 10, "citations_str": 40, "citation_hit": 10,
            "latency_retrieve_ms": 12, "latency_rerank_ms": 12,
            "latency_context_ms": 12, "latency_llm_ms": 12, "latency_total_ms": 12,
            "error": 30, "status": 10,
        }
        for col_idx, col_name in enumerate(COLUMN_ORDER, 1):
            width = col_widths.get(col_name, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze panes (header row + first 3 columns)
        ws.freeze_panes = "D2"

    def _write_stats_sheet(self, ws):
        """Ghi sheet thống kê tổng hợp."""
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        n = len(self.entries)
        if n == 0:
            ws.cell(row=1, column=1, value="Chưa có dữ liệu")
            return

        # Title
        ws.cell(row=1, column=1, value="THỐNG KÊ PIPELINE").font = title_font
        ws.cell(row=2, column=1, value=f"Run: {self.run_id}")
        ws.cell(row=3, column=1, value=f"Tổng queries: {n}")
        ws.cell(row=4, column=1, value=f"Thời gian: {self.entries[0].timestamp} → {self.entries[-1].timestamp}")

        # ─── Tier distribution ───
        row = 6
        ws.cell(row=row, column=1, value="PHÂN BỐ TIER").font = header_font
        row += 1
        for h_idx, h in enumerate(["Tier", "Số lượng", "Tỷ lệ"], 1):
            c = ws.cell(row=row, column=h_idx, value=h)
            c.font = header_font
            c.fill = header_fill

        tier_counts = {}
        for e in self.entries:
            tier_counts[e.tier] = tier_counts.get(e.tier, 0) + 1
        row += 1
        for tier_name in ["local", "api", "none"]:
            count = tier_counts.get(tier_name, 0)
            ws.cell(row=row, column=1, value=tier_name.upper())
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/n:.1%}")
            if tier_name in TIER_COLORS:
                ws.cell(row=row, column=1).fill = PatternFill(
                    start_color=TIER_COLORS[tier_name],
                    end_color=TIER_COLORS[tier_name],
                    fill_type="solid"
                )
            row += 1

        # ─── Decision distribution ───
        row += 1
        ws.cell(row=row, column=1, value="PHÂN BỐ GATING DECISION").font = header_font
        row += 1
        for h_idx, h in enumerate(["Decision", "Số lượng", "Tỷ lệ"], 1):
            c = ws.cell(row=row, column=h_idx, value=h)
            c.font = header_font
            c.fill = header_fill

        dec_counts = {}
        for e in self.entries:
            dec_counts[e.gating_decision] = dec_counts.get(e.gating_decision, 0) + 1
        row += 1
        for dec in ["answer", "cautious", "abstain", "ask_back"]:
            count = dec_counts.get(dec, 0)
            ws.cell(row=row, column=1, value=dec.upper())
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/n:.1%}")
            row += 1

        # ─── Latency stats ───
        row += 1
        ws.cell(row=row, column=1, value="THỐNG KÊ LATENCY (ms)").font = header_font
        row += 1
        for h_idx, h in enumerate(["Bước", "Trung bình", "Min", "Max"], 1):
            c = ws.cell(row=row, column=h_idx, value=h)
            c.font = header_font
            c.fill = header_fill

        latency_fields = [
            ("Retrieve", "latency_retrieve_ms"),
            ("Rerank", "latency_rerank_ms"),
            ("Context Build", "latency_context_ms"),
            ("LLM", "latency_llm_ms"),
            ("TOTAL", "latency_total_ms"),
        ]
        row += 1
        for label, field_name in latency_fields:
            vals = [getattr(e, field_name) for e in self.entries if getattr(e, field_name) > 0]
            if vals:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=round(sum(vals)/len(vals), 1))
                ws.cell(row=row, column=3, value=round(min(vals), 1))
                ws.cell(row=row, column=4, value=round(max(vals), 1))
                row += 1

        # ─── Error rate ───
        row += 1
        error_count = sum(1 for e in self.entries if e.error)
        ws.cell(row=row, column=1, value="Tỷ lệ lỗi").font = header_font
        ws.cell(row=row, column=2, value=f"{error_count}/{n} ({error_count/n:.1%})")

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _write_tier_sheet(self, ws):
        """Ghi sheet phân tích tier chi tiết."""
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        n = len(self.entries)
        if n == 0:
            ws.cell(row=1, column=1, value="Chưa có dữ liệu")
            return

        ws.cell(row=1, column=1, value="PHÂN TÍCH 2-TIER ROUTING").font = Font(bold=True, size=12)

        # Tier breakdown with score ranges
        row = 3
        headers = ["Tier", "Count", "%", "Avg Score Top1", "Avg Margin", "Avg Latency (ms)", "Errors"]
        for h_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=h_idx, value=h)
            c.font = header_font
            c.fill = header_fill

        for tier_name in ["local", "api", "none"]:
            row += 1
            tier_entries = [e for e in self.entries if e.tier == tier_name]
            count = len(tier_entries)
            if count == 0:
                ws.cell(row=row, column=1, value=tier_name.upper())
                ws.cell(row=row, column=2, value=0)
                continue

            avg_score = sum(e.score_rerank_top1 for e in tier_entries) / count
            avg_margin = sum(e.margin for e in tier_entries) / count
            avg_latency = sum(e.latency_total_ms for e in tier_entries) / count
            errors = sum(1 for e in tier_entries if e.error)

            ws.cell(row=row, column=1, value=tier_name.upper())
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/n:.1%}")
            ws.cell(row=row, column=4, value=round(avg_score, 2))
            ws.cell(row=row, column=5, value=round(avg_margin, 2))
            ws.cell(row=row, column=6, value=round(avg_latency, 1))
            ws.cell(row=row, column=7, value=errors)

            if tier_name in TIER_COLORS:
                ws.cell(row=row, column=1).fill = PatternFill(
                    start_color=TIER_COLORS[tier_name],
                    end_color=TIER_COLORS[tier_name],
                    fill_type="solid"
                )

        # Cost analysis
        row += 2
        ws.cell(row=row, column=1, value="PHÂN TÍCH CHI PHÍ").font = Font(bold=True, size=12)
        row += 1
        local_count = sum(1 for e in self.entries if e.tier == "local")
        api_count = sum(1 for e in self.entries if e.tier == "api")
        none_count = sum(1 for e in self.entries if e.tier == "none")

        ws.cell(row=row, column=1, value="Queries xử lý LOCAL (miễn phí):")
        ws.cell(row=row, column=2, value=f"{local_count} ({local_count/n:.1%})" if n else "0")
        row += 1
        ws.cell(row=row, column=1, value="Queries gọi API (tốn phí):")
        ws.cell(row=row, column=2, value=f"{api_count} ({api_count/n:.1%})" if n else "0")
        row += 1
        ws.cell(row=row, column=1, value="Queries KHÔNG gọi LLM (tiết kiệm):")
        ws.cell(row=row, column=2, value=f"{none_count} ({none_count/n:.1%})" if n else "0")
        row += 1
        savings = (local_count + none_count) / n * 100 if n else 0
        ws.cell(row=row, column=1, value="Tỷ lệ tiết kiệm API:").font = Font(bold=True, size=11)
        ws.cell(row=row, column=2, value=f"{savings:.1f}%").font = Font(bold=True, size=11, color="008000")

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def get_summary(self, include_prior: bool = False) -> Dict[str, Any]:
        """Trả về summary stats. include_prior=True gộp cả entries cũ (append mode)."""
        entries = (self._prior_entries + self.entries) if (include_prior and self.append) else self.entries
        n = len(entries)
        if n == 0:
            return {"total": 0}

        tier_counts = {}
        dec_counts = {}
        for e in entries:
            tier_counts[e.tier] = tier_counts.get(e.tier, 0) + 1
            dec_counts[e.gating_decision] = dec_counts.get(e.gating_decision, 0) + 1

        return {
            "total": n,
            "run_id": self.run_id,
            "tiers": {k: {"count": v, "rate": round(v/n, 4)} for k, v in tier_counts.items()},
            "decisions": {k: {"count": v, "rate": round(v/n, 4)} for k, v in dec_counts.items()},
            "avg_latency_total_ms": round(sum(e.latency_total_ms for e in entries) / n, 1),
            "error_rate": round(sum(1 for e in entries if e.error) / n, 4),
            "api_savings_rate": round(
                sum(1 for e in entries if e.tier in ("local", "none")) / n, 4
            ),
        }

    def print_summary(self):
        """In summary ra console (chỉ entries run hiện tại)."""
        s = self.get_summary(include_prior=False)
        if s["total"] == 0:
            print("[Logger] No entries")
            return

        print("\n" + "=" * 60)
        print(f"  PIPELINE LOG SUMMARY — {s['run_id']}")
        print("=" * 60)
        print(f"  Total queries:  {s['total']}")
        print(f"  Avg latency:    {s['avg_latency_total_ms']:.0f} ms")
        print(f"  Error rate:     {s['error_rate']:.1%}")
        print()
        print("  ┌─ TIER DISTRIBUTION ──────────────────┐")
        for tier in ["local", "api", "none"]:
            info = s["tiers"].get(tier, {"count": 0, "rate": 0})
            bar = "█" * int(info["rate"] * 30)
            label = {"local": "LOCAL (free) ", "api": "API (paid)   ", "none": "NONE (skip)  "}
            print(f"  │ {label.get(tier, tier):14s} {info['count']:4d}  {info['rate']:6.1%} {bar}")
        print(f"  │ API savings: {s['api_savings_rate']:.1%}")
        print("  └────────────────────────────────────────┘")
        print()
        print("  ┌─ GATING DECISIONS ────────────────────┐")
        for dec in ["answer", "cautious", "abstain", "ask_back"]:
            info = s["decisions"].get(dec, {"count": 0, "rate": 0})
            bar = "█" * int(info["rate"] * 30)
            print(f"  │ {dec:12s} {info['count']:4d}  {info['rate']:6.1%} {bar}")
        print("  └────────────────────────────────────────┘")
        print()

    def close(self):
        """Flush cuối cùng và đóng logger."""
        self.flush_excel()
        self.print_summary()
        print(f"[Logger] Files saved:")
        print(f"  Excel:  {self.xlsx_path}")
        print(f"  CSV:    {self.csv_path}")
        print(f"  JSONL:  {self.jsonl_path}")
